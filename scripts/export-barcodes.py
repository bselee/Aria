"""
Export barcodes from OfficialBarcodes.xlsx to TypeScript registry.

Usage:
    python scripts/export-barcodes.py [path-to-xlsx]

Outputs to stdout — redirect to update barcode-registry.ts:
    python scripts/export-barcodes.py > src/lib/dash/barcode-registry-export.ts
"""

import openpyxl, json, sys

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else 'OfficialBarcodes.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

records = {}
for r in range(2, ws.max_row + 1):
    pid = str(ws.cell(r, 1).value or '').strip()
    upc = ws.cell(r, 4).value
    if upc:
        upc_str = str(upc).replace(' ', '').replace('-', '')
        records[pid] = upc_str

print('// Auto-generated from OfficialBarcodes.xlsx')
print(f'// Source: {xlsx_path}')
print('// Run: python scripts/export-barcodes.py')
print()
print('export const FULL_BARCODE_REGISTRY: Record<string, string> = {')
for pid in sorted(records):
    print(f"  '{pid}': '{records[pid]}',")
print('};')
print(f'// Total: {len(records)} entries with barcodes')
